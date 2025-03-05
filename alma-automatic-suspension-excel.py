import openpyxl
import time
import os
from datetime import date

# The date.
current_date = date.today()

# Directory URL which will be assigned from user-input.
# Initials to go at the end of the suspension note, also assigned from user-input.
# outputFilePath refers to the user's home, on windows it is the "Documents" folder and creates a new text file named after the date.
url = ""
initials = ""
output_file_path = os.path.expanduser("~")

# This makes the Alma-Automatic-Suspensions folder
# noinspection PyBroadException
try:
    os.makedirs(os.path.expanduser("~") + "/Alma-Automatic-Suspensions")
    output_file_path = output_file_path + "/Alma-Automatic-Suspensions"
    print(f"Alma Automatic Suspension Directory created in : {output_file_path}")
except:
    print(f"Alma Automatic Suspension Directory found in: {output_file_path}")

output_file_path = output_file_path + "/alma-automatic-suspension-output-" + str(current_date.month) + "-" + str(current_date.day) + "-" + str(current_date.year) + ".txt"

# These variables are in regards to legal letters. BreakLine is the line to seperate the legal letter suspensions.
# legalLetterRequirement is the count of days an item can be overdue before being sent a legal letter, if it's one day over it gets seperated.
break_line = False
legal_letter_requirement = 30

# This is a loop to confirm your initials to display near the end of the suspension note.
while initials == "":
    newInitials = input("Please insert your initials: ")
    confirmation = input("The format for the note will be: -" + str(newInitials) + "  | Is this correct? Y or N \n")

    if str.lower(confirmation) == "y":
        initials = newInitials

# This is a loop to get a working directory url to the excel sheet. If the format has backslashes, it will be convertered for use.
while url == "":
    newUrl = input("Please drag the file to this window, and make sure this window is selected, then press enter:\n")
    newUrl = newUrl.replace("\"", "")
    newUrl = newUrl.replace("\\", "/")

    print("Filepath recieved: " + newUrl)
    if os.path.exists(newUrl):
        url = newUrl
    else:
        print("File not found.")

# OpenPyXL is a python library to read/write Excel files.
# It runs locally and makes know pulls or requests to send data through the internet and is completely safe to use.
# It is a little bit out of date but can work for our purpose since the data is also in an old format. (2010 Excel Files)
workbook = openpyxl.load_workbook(url)
active_sheet = workbook.active

data = {}

'''
data structure
data
    - userId 
        -- row
        -- days overdue
        -- item#
            --- item title 
            --- process status 
'''
previous_user_id = None # placeholder
previous_iterator = 0

# Adds to the dictionary where the key is the userId, and the values are another dictionary. 
def add_data(user_id):
    global previous_user_id
    global previous_iterator
    # Assigning a new user Id.
    data[user_id] = {}
    data[user_id]["Items"] = {}
    data[user_id]["Row"] = row
    data[user_id]["DaysOverdue"] = active_sheet["G"][row].value
    data[user_id]["Name"] = active_sheet["C"][row].value + ", " + active_sheet["B"][row].value
    data[user_id]["Items"]["Item1"] = {
        "Title": active_sheet["I"][row].value,
        "ProcessStatus": active_sheet["J"][row].value,
        "Barcode": active_sheet["H"][row].value
    }

    # Assigning previous variables to find multiple items in the sheet.
    previous_user_id = user_id
    previous_iterator = 1
    print("[AAS] New UserId Found: " + str(user_id))


# This section organizes it into data.
for row in range(0, active_sheet.max_row):

    # Iterates over sheet to find the user id rows and to not cause dupliates with multiple lost items.
    for col in active_sheet.iter_cols(1, 1):
        cell_user_id = col[row].value

        # A blank entry was listed which in most cases is another item, this assigns new items to the previous id.
        if cell_user_id == None:
            if previous_user_id != None:
                previous_iterator += 1
                print("[AAS] Another Item Found, Item " + str(previous_iterator))

                # This if statement checks to see if the new item is more recently overdue to sort the id higher later on the list.
                daysOverdue = active_sheet["G"][row].value
                if data[previous_user_id]["DaysOverdue"] < daysOverdue:
                       data[previous_user_id]["DaysOverdue"] = daysOverdue

                data[previous_user_id]["Items"]["Item" + str(previous_iterator)] = {
                "Title": active_sheet["I"][row].value,
                "ProcessStatus": active_sheet["J"][row].value,
                "Barcode": active_sheet["H"][row].value
            }
            continue
        # This checks to see if the id is a string since Community Member's id's start with a letter.
        if isinstance(cell_user_id, str):
            if cell_user_id.isnumeric():
                add_data(cell_user_id)
            else:
                # Community Member Id
                if cell_user_id[1:].isnumeric():
                    add_data(cell_user_id)
        else:
            add_data(cell_user_id)

# Sorts data using lambda to the DaysOverdue from earliest to longest overdues.
sorted_data = dict(sorted(data.items(), key=lambda x: x[1]['DaysOverdue']))

# Opens the output file path to start writing in the utf-8 encoding format.
with open(output_file_path, "w", encoding='utf-8') as file:
    for cell_user_id in sorted_data:
        # This is the data produced by a person.

        suspensionNote = "SUSPENDED / Instance#X / LOST ["
        itemsLost = ""
        itemBarcodes = ""
        # This assigns itemsLost and itemBarcodes as a list on a string for the text file.
        for item in data[cell_user_id]["Items"]:
            if data[cell_user_id]["Items"][item]["ProcessStatus"] == "LOST":
                        itemsLost = itemsLost + "\'" + data[cell_user_id]["Items"][item]["Title"] + "\',"
                        itemBarcodes = itemBarcodes + str(data[cell_user_id]["Items"][item]["Barcode"]) + ","

        if itemsLost == "":
            # This person does not match the requirements to be suspended, their items are overdue but not lost.
            continue
        else:
            # These cut off the last comma from the lists.
            itemsLost = itemsLost[:len(itemsLost)-1]
            itemBarcodes = itemBarcodes[:len(itemBarcodes)-1]

            suspensionNote = suspensionNote + itemsLost + "]-unresolved- " + str(current_date.month) + "/" + str(current_date.day) + "/" + str(current_date.year) + "-" + initials
            # Legal letter.
            if data[cell_user_id]["DaysOverdue"] > legal_letter_requirement and break_line == False:
                break_line = True
                file.write("\n\n\n\nThe following are people who are eligible to receive a legal letter. Fees are not taken into an account int this list. Requirement: " + str(legal_letter_requirement) + " days overdue.\n\n\n\n")
                 
            ''' Format
            UserId:
            Name: Last, First
            X Days Overdue.
            Suspension note. (You can triple click this line in the text file to easily copy it.)
            Item barcodes in order of the list on the suspension note.
            '''
            file.write("UserId: " + str(cell_user_id) + "\nName: " + str(data[cell_user_id]["Name"]) + "\n" + str(data[cell_user_id]["DaysOverdue"]) + " days overdue.\n" + suspensionNote + "\nThese are the item barcodes in order: " + itemBarcodes + "\n")

# This opens up the exported text file.
# noinspection PyBroadException
try:
    os.system(output_file_path)
except Exception as e:
    # This has never happened during testing but is a precaution just in case something goes wrong.
    print("An error occured while opening the notepad, but can still be found under the Documents folder.")

# Extends the program's lifetime to view the window just in case it's needed.
print("This window will close in 30 seconds.")
time.sleep(30)